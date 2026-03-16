import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import calendar
import os

# --- Настройки ---
input_file = "Календарь Событий Октябрь 2025.xlsx"  # исходный файл
output_file = "Календарь Событий Ноябрь 2025.xlsx"

# Открываем книгу
wb = openpyxl.load_workbook(input_file)
ws1 = wb.active

# Проверка и получение года
year_cell = ws1["A1"].value
if isinstance(year_cell, int):
    year = year_cell
else:
    try:
        year = int(str(year_cell).strip())
    except:
        raise ValueError(f"Не удалось определить год из A1: {year_cell}")

# Получаем месяц из A2, но будем подставлять следующий месяц
month_cell = ws1["A2"].value
months_ru = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
}
month_map_ru = {v: k for k, v in months_ru.items()}

# Определяем номер месяца (если не удалось, берём октябрь по умолчанию)
try:
    old_month_num = month_map_ru[str(month_cell).strip()]
except:
    old_month_num = 10  # Октябрь

# Новый месяц
new_month_num = old_month_num + 1 if old_month_num < 12 else 1
new_year = year if old_month_num < 12 else year + 1

# --- Обновляем первую страницу ---
ws1["A1"].value = new_year
ws1["A2"].value = months_ru[new_month_num]

# Удаляем старые даты (начиная с ряда 4)
for row in ws1.iter_rows(min_row=4):
    for cell in row:
        cell.value = None
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Заполняем сетку дат
cal = calendar.Calendar(firstweekday=0)  # неделя с понедельника
month_days = cal.monthdatescalendar(new_year, new_month_num)
start_row = 4

for week in month_days:
    for col_idx, date in enumerate(week, start=1):
        if date.month == new_month_num:
            ws1.cell(row=start_row, column=col_idx, value=date.strftime("%d.%m.%Y"))
        else:
            ws1.cell(row=start_row, column=col_idx, value=None)
        ws1.cell(row=start_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
    start_row += 1

# --- Очищаем вторую страницу ---
if len(wb.sheetnames) > 1:
    ws2 = wb.worksheets[1]

    # Разъединяем все объединённые ячейки
    for merged in list(ws2.merged_cells.ranges):
        ws2.unmerge_cells(str(merged))

    # Очищаем содержимое
    for row in ws2.iter_rows():
        for cell in row:
            cell.value = None
            cell.alignment = Alignment(horizontal="center", vertical="center")
else:
    # Если второй страницы нет, создаём
    ws2 = wb.create_sheet("Матчи")

# --- Сохраняем ---
wb.save(output_file)
print(f"Файл обновлён и сохранён как {output_file}")
