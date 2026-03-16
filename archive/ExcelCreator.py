import calendar
import csv
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

# ---- Определяем следующий месяц ----
today = datetime.today()
if today.month == 12:
    year_next = today.year + 1
    month_next = 1
else:
    year_next = today.year
    month_next = today.month + 1

month_english = calendar.month_name[month_next].lower()  # английское название для файлов

# Пути
csv_dir = rf"C:\Users\efim\PycharmProjects\PythonProject2\filtered_{month_english}_{year_next}"
csv_path = os.path.join(csv_dir, f"filtered_matches_{month_english}_{year_next}.csv")
tournaments_dir = "tournaments"

# Загружаем список турниров для фильтрации
def load_tournaments():
    tournaments_set = set()
    if not os.path.exists(tournaments_dir):
        print(f"Папка с турнирами не найдена: {tournaments_dir}")
        return tournaments_set
    for fname in os.listdir(tournaments_dir):
        if fname.lower().endswith(".txt"):
            with open(os.path.join(tournaments_dir, fname), encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line:
                        tournaments_set.add(line)
    return tournaments_set

tournaments_set = load_tournaments()

# ---- Создание книги ----
wb = Workbook()

# ---- Первая вкладка: календарь ----
months_ru = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
}
days_ru = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

ws1 = wb.active
ws1.title = "Календарь"

# Ширина колонок
for col in range(1, 8):
    ws1.column_dimensions[chr(64 + col)].width = 5

# Стили
center = Alignment(horizontal="center", vertical="center")
bold_font = Font(bold=True, size=12)
gray_fill = PatternFill("solid", fgColor="DDDDDD")

# Заголовки года и месяца
ws1.merge_cells("A1:G1")
ws1["A1"] = str(year_next)
ws1["A1"].font = Font(bold=True, size=14)
ws1["A1"].alignment = center

ws1.merge_cells("A2:G2")
ws1["A2"] = months_ru[month_next]
ws1["A2"].font = Font(bold=True, size=13)
ws1["A2"].alignment = center

# Дни недели
for i, day in enumerate(days_ru, start=1):
    ws1.cell(row=3, column=i, value=day).font = bold_font
    ws1.cell(row=3, column=i).alignment = center

# Календарная сетка
cal = calendar.Calendar(firstweekday=0)
month_days = cal.monthdatescalendar(year_next, month_next)
row = 4
for week in month_days:
    for col, date in enumerate(week, start=1):
        cell = ws1.cell(row=row, column=col, value=date.day)
        cell.alignment = center
        if date.month != month_next:
            cell.fill = gray_fill
    row += 1

# ---- Вторая вкладка: матчи ----
ws2 = wb.create_sheet(title="Матчи")
headers = ["Турнир", "Дата", "Время", "Участники"]
for col, h in enumerate(headers, start=1):
    ws2.cell(row=1, column=col, value=h).font = bold_font
    ws2.cell(row=1, column=col).alignment = center

# Чтение CSV и фильтрация
if not os.path.exists(csv_path):
    print(f"CSV файл с матчами не найден: {csv_path}")
else:
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        row_idx = 2
        for row_data in reader:
            tournament = row_data.get("tournament", "")
            # Жёсткая фильтрация по ключу: если ключ из tournaments_set содержится в названии турнира
            if any(key in tournament for key in tournaments_set):
                ws2.cell(row=row_idx, column=1, value=tournament).alignment = Alignment(wrap_text=True)
                ws2.cell(row=row_idx, column=2, value=row_data.get("date", ""))
                ws2.cell(row=row_idx, column=3, value=row_data.get("time", ""))
                ws2.cell(row=row_idx, column=4, value=row_data.get("participants", "")).alignment = Alignment(wrap_text=True)
                row_idx += 1

# Сохранение Excel
output_dir = "output_excels"
os.makedirs(output_dir, exist_ok=True)
output_filename = os.path.join(output_dir, f"calendar_with_matches_{month_english}_{year_next}.xlsx")
wb.save(output_filename)
print(f"Excel файл успешно создан: {output_filename}")
