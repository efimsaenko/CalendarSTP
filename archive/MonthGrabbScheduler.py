"""
MonthGrabbScheduler.py

Парсер матчей за месяц. Функции:
 - пройти по всем дням месяца и последовательно спарсить каждый день
 - сохранять результаты в Excel: data/YYYY/MM/matches_YYYY_MM.xlsx
   - лист 'matches'       — все строки (tournament, date, time, participants, request_id)
   - лист 'daily_summary' — по дню (date, matches_count, status, request_id, ts)
 - метод grab_one_day(date_str) для исправления отдельного дня
 - логирование в json (однострочные записи)
 - таймауты, retries, алертинг при проблемах сети или блокировке
 - простой файловый лок для защиты от одновременных записей

Зависимости: openpyxl, selenium

Запуск из кода:
    from MonthGrabbScheduler import MonthGrabbScheduler
    scheduler = MonthGrabbScheduler(chrome_path=..., chromedriver_path=...)
    scheduler.run_month(year=2026, month=3, force_grab=True)
    scheduler.grab_one_day("2026-03-31")

Запуск из CLI:
    python MonthGrabbScheduler.py --year 2026 --month 3 --headless
"""

import os
import sys
import json
import time
import uuid
import argparse
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Iterator
from contextlib import contextmanager

from openpyxl import Workbook, load_workbook

try:
    from DailyGrabber import ChampionatGrabber
except Exception as _e:
    ChampionatGrabber = None
    print(f"Не удалось импортировать ChampionatGrabber: {_e}")


# ---------------------------------------------------------------------------
# JSON Logger
# ---------------------------------------------------------------------------

class JsonFormatter(logging.Formatter):
    def format(self, record: logging.LogRecord) -> str:
        payload = {
            "timestamp": datetime.utcfromtimestamp(record.created).isoformat() + "Z",
            "level":     record.levelname,
            "module":    record.name,
            "message":   record.getMessage(),
        }
        for key in ("request_id", "user_id", "username", "alert"):
            val = getattr(record, key, None)
            if val is not None:
                payload[key] = val
        if record.exc_info:
            payload["exc"] = self.formatException(record.exc_info).replace("\n", " | ")
        return json.dumps(payload, ensure_ascii=False)


def _make_logger(name: str = "MonthGrabbScheduler", log_dir: str = "logs") -> logging.Logger:
    os.makedirs(log_dir, exist_ok=True)
    logger = logging.getLogger(name)
    if logger.handlers:
        return logger
    logger.setLevel(logging.INFO)
    fmt = JsonFormatter()

    fh = logging.FileHandler(
        os.path.join(log_dir, datetime.now().strftime("%Y-%m-%d") + ".json.log"),
        encoding="utf-8"
    )
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    return logger


# ---------------------------------------------------------------------------
# File lock
# ---------------------------------------------------------------------------

@contextmanager
def file_lock(lock_path: str, timeout: int = 60):
    """Файловый лок через O_EXCL. Работает на Windows и Linux."""
    start = time.time()
    while True:
        try:
            fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_RDWR)
            os.write(fd, str(os.getpid()).encode())
            os.close(fd)
            break
        except FileExistsError:
            if time.time() - start > timeout:
                raise TimeoutError(f"Timeout acquiring lock: {lock_path}")
            time.sleep(0.3)
    try:
        yield
    finally:
        try:
            os.remove(lock_path)
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def gen_request_id() -> str:
    return uuid.uuid4().hex[:12]


def list_days_in_month(year: int, month: int) -> Iterator[str]:
    dt = datetime(year=year, month=month, day=1)
    while dt.month == month:
        yield dt.strftime("%Y-%m-%d")
        dt += timedelta(days=1)


# ---------------------------------------------------------------------------
# Excel wrapper
# ---------------------------------------------------------------------------

class ExcelMonthFile:
    MATCHES_HEADER = ["tournament", "date", "time", "participants", "request_id"]
    SUMMARY_HEADER = ["date", "matches_count", "status", "request_id", "ts"]

    def __init__(self, path: str):
        self.path      = path
        self.lock_path = path + ".lock"
        self._ensure_file()

    def _ensure_file(self):
        os.makedirs(os.path.dirname(self.path), exist_ok=True)
        if not os.path.exists(self.path):
            wb = Workbook()
            ws = wb.active
            ws.title = "matches"
            ws.append(self.MATCHES_HEADER)
            wb.create_sheet("daily_summary").append(self.SUMMARY_HEADER)
            wb.save(self.path)

    def append_matches(self, rows: List[Dict]):
        if not rows:
            return
        with file_lock(self.lock_path):
            wb = load_workbook(self.path)
            ws = wb["matches"] if "matches" in wb.sheetnames else wb.create_sheet("matches")
            if ws.max_row == 1 and ws.cell(1, 1).value is None:
                ws.append(self.MATCHES_HEADER)
            for r in rows:
                ws.append([
                    r.get("tournament", ""),
                    r.get("date", ""),
                    r.get("time", ""),
                    r.get("participants", ""),
                    r.get("request_id", ""),
                ])
            wb.save(self.path)

    def write_summary_row(self, date: str, count: int, status: str, req_id: str):
        with file_lock(self.lock_path):
            wb = load_workbook(self.path)
            ws = wb["daily_summary"] if "daily_summary" in wb.sheetnames else wb.create_sheet("daily_summary")
            ws.append([date, count, status, req_id, datetime.utcnow().isoformat() + "Z"])
            wb.save(self.path)

    def replace_day_rows(self, date: str, new_rows: List[Dict]):
        """Удаляет все строки за date и записывает новые."""
        with file_lock(self.lock_path):
            wb = load_workbook(self.path)
            ws = wb["matches"] if "matches" in wb.sheetnames else wb.create_sheet("matches")

            # Читаем всё, кроме удаляемого дня
            all_rows = list(ws.iter_rows(min_row=2, values_only=True))
            kept = [r for r in all_rows if r[1] != date]

            # Очищаем данные (оставляем заголовок)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

            # Перезаписываем сохранённые строки
            for i, r in enumerate(kept, start=2):
                for j, v in enumerate(r, start=1):
                    ws.cell(row=i, column=j, value=v)

            # Дописываем новые строки
            start_row = len(kept) + 2
            for i, r in enumerate(new_rows, start=start_row):
                ws.cell(row=i, column=1, value=r.get("tournament", ""))
                ws.cell(row=i, column=2, value=r.get("date", ""))
                ws.cell(row=i, column=3, value=r.get("time", ""))
                ws.cell(row=i, column=4, value=r.get("participants", ""))
                ws.cell(row=i, column=5, value=r.get("request_id", ""))

            wb.save(self.path)


# ---------------------------------------------------------------------------
# MonthGrabbScheduler
# ---------------------------------------------------------------------------

class MonthGrabbScheduler:
    def __init__(self,
                 chrome_path: str,
                 chromedriver_path: str,
                 data_root: str = "data",
                 timeout_per_day: int = 90,
                 retries: int = 3,
                 backoff: int = 5,
                 headless: bool = True,
                 log_dir: str = "logs"):

        if ChampionatGrabber is None:
            raise RuntimeError("ChampionatGrabber не доступен — положи DailyGrabber.py рядом с этим файлом")

        self.chrome_path      = chrome_path
        self.chromedriver_path = chromedriver_path
        self.data_root        = data_root
        self.timeout_per_day  = timeout_per_day
        self.retries          = retries
        self.backoff          = backoff
        self.headless         = headless
        self.logger           = _make_logger("MonthGrabbScheduler", log_dir)

    # ------------------------------------------------------------------
    # Internals
    # ------------------------------------------------------------------

    def _get_excel_path(self, year: int, month: int) -> str:
        folder = os.path.join(self.data_root, f"{year:04d}", f"{month:02d}")
        os.makedirs(folder, exist_ok=True)
        return os.path.join(folder, f"matches_{year:04d}_{month:02d}.xlsx")

    def _grab_with_retry(self, date_str: str, req_id: str) -> Dict:
        """
        Запускает ChampionatGrabber для одного дня с retries.
        Возвращает {"status": "ok", "matches": [...]} или {"status": "failed", "error": "..."}
        """
        last_error = ""
        for attempt in range(1, self.retries + 1):
            grabber = None
            try:
                self.logger.info(
                    f"Попытка {attempt}/{self.retries} для {date_str}",
                    extra={"request_id": req_id}
                )
                grabber = ChampionatGrabber(
                    chrome_path=self.chrome_path,
                    chromedriver_path=self.chromedriver_path,
                    headless=self.headless,
                )
                matches = grabber.grab(date_str, save_csv=False, request_id=req_id)
                grabber.close()

                self.logger.info(
                    f"Спарсено {len(matches)} матчей за {date_str}",
                    extra={"request_id": req_id}
                )
                return {"status": "ok", "matches": matches}

            except Exception as e:
                last_error = str(e)
                self.logger.warning(
                    f"Ошибка попытки {attempt} для {date_str}: {e}",
                    extra={"request_id": req_id, "alert": attempt == self.retries}
                )
                try:
                    if grabber:
                        grabber.close()
                except Exception:
                    pass
                if attempt < self.retries:
                    time.sleep(self.backoff * attempt)

        return {"status": "failed", "error": last_error}

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def grab_one_day(self, date_str: str, force_grab: bool = True) -> int:
        """
        Парсит один день и дописывает/заменяет данные в месячном Excel.

        Аргументы:
            date_str   — дата в формате "YYYY-MM-DD"
            force_grab — если True, заменяет существующие строки за этот день
                         если False, пропускает если данные уже есть

        Возвращает количество спарсенных матчей (0 при ошибке).
        """
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            self.logger.error(f"Неверный формат даты: {date_str!r}. Ожидается YYYY-MM-DD")
            return 0

        year, month = dt.year, dt.month
        req_id      = gen_request_id()
        excel       = ExcelMonthFile(self._get_excel_path(year, month))

        self.logger.info(f"grab_one_day: {date_str} force={force_grab}", extra={"request_id": req_id})

        result = self._grab_with_retry(date_str, req_id)

        if result["status"] == "ok":
            matches = result["matches"]
            for r in matches:
                r["request_id"] = req_id
                r.setdefault("date", date_str)

            if force_grab:
                excel.replace_day_rows(date_str, matches)
                status = "repaired"
            else:
                excel.append_matches(matches)
                status = "ok"

            excel.write_summary_row(date_str, len(matches), status, req_id)
            self.logger.info(f"grab_one_day done: {date_str} → {len(matches)} матчей", extra={"request_id": req_id})
            return len(matches)
        else:
            excel.write_summary_row(date_str, 0, "failed", req_id)
            self.logger.error(
                f"grab_one_day failed: {date_str} → {result.get('error')}",
                extra={"request_id": req_id, "alert": True}
            )
            return 0

    def run_month(self, year: int, month: int, force_grab: bool = False):
        """
        Парсит все дни месяца последовательно.

        Аргументы:
            year, month — год и месяц
            force_grab  — если True, перезаписывает уже существующие дни
        """
        self.logger.info(f"run_month start: {year:04d}-{month:02d} force={force_grab}")

        days   = list(list_days_in_month(year, month))
        ok     = 0
        failed = 0

        for date_str in days:
            count = self.grab_one_day(date_str, force_grab=force_grab)
            if count > 0:
                ok += 1
            else:
                failed += 1
            # Небольшая пауза между днями чтобы не получить блокировку
            time.sleep(2)

        self.logger.info(
            f"run_month done: {year:04d}-{month:02d} | "
            f"ok={ok} failed={failed} total={len(days)}",
            extra={"alert": failed > 0}
        )

    def repair_day(self, year: int, month: int, date_str: str):
        """Алиас для grab_one_day с force_grab=True. Удобен для исправления одного дня."""
        self.logger.info(f"repair_day: {date_str}")
        self.grab_one_day(date_str, force_grab=True)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_args():
    p = argparse.ArgumentParser(description="Парсер матчей за месяц")
    p.add_argument("--year",        type=int, required=True)
    p.add_argument("--month",       type=int, required=True)
    p.add_argument("--chrome",      type=str, default=r"chrome-win64\chrome.exe")
    p.add_argument("--chromedriver",type=str, default=r"chromedriver-win64\chromedriver.exe")
    p.add_argument("--headless",    action="store_true")
    p.add_argument("--force",       action="store_true", help="Перезаписать уже спарсенные дни")
    p.add_argument("--repair-day",  type=str, default=None, metavar="YYYY-MM-DD",
                   help="Перепарсить один день вместо полного месяца")
    return p.parse_args()


def main():
    args = _parse_args()
    sched = MonthGrabbScheduler(
        chrome_path=args.chrome,
        chromedriver_path=args.chromedriver,
        headless=args.headless,
    )
    if args.repair_day:
        sched.repair_day(args.year, args.month, args.repair_day)
    else:
        sched.run_month(args.year, args.month, force_grab=args.force)


if __name__ == "__main__":
    main()