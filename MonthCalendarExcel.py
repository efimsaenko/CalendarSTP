# MonthCalendarExcel.py
"""
Создаёт Excel-календарь на основе шаблона (март 2026).

Лист 1 «Календарь спорт. активности»:
  - A2 = год, B2 = месяц
  - Строки 4..15: пары (дата / контент) по 2 строки на неделю
  - Покраска по status.csv, контент = короткие названия турниров дня

Лист 2 «Детально»:
  - Строка 1 — заголовок (не трогаем)
  - Со строки 2 — матчи из MM.YYYY_filtered.xlsx
  - Колонки: Турнир(кратко) | Дата | Время | Участвующие команды

Публичный API:
  create_month_calendar(month_str, base_dir, template_path) -> str
  get_day_cell_map(ws, year, month) -> dict
  write_day_content(ws, cell_map, day, text)

Запуск:
  python MonthCalendarExcel.py 03.2026
"""

import os, sys, re, shutil, logging
from datetime import date, datetime, timedelta
from typing import Dict, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

try:
    from logger import get_logger
    logger = get_logger("MonthCalendarExcel")
except ImportError:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    logger = logging.getLogger("MonthCalendarExcel")

# ---------------------------------------------------------------------------
# Константы
# ---------------------------------------------------------------------------

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Календарь_Событий.xlsx")

MONTH_NAMES_RU = {
    1:"Январь", 2:"Февраль", 3:"Март",    4:"Апрель",
    5:"Май",    6:"Июнь",    7:"Июль",    8:"Август",
    9:"Сентябрь", 10:"Октябрь", 11:"Ноябрь", 12:"Декабрь",
}

STATUS_COLORS = {"red": "FFFF0000", "yellow": "FFFFFF00", "green": "FF92D050"}
DEFAULT_COLOR = "FFFFFFFF"
GREY_FILL     = PatternFill(fill_type="solid", fgColor="FFD9D9D9")

WEEK_ROW_PAIRS = [(4,5),(6,7),(8,9),(10,11),(12,13),(14,15)]
COL_LETTERS    = {0:"A",1:"B",2:"C",3:"D",4:"E",5:"F",6:"G"}

DETAIL_COL_WIDTHS = {"A":36, "B":15, "C":13, "D":52}

# ---------------------------------------------------------------------------
# Пути
# ---------------------------------------------------------------------------

def _base(base_dir=None):
    return base_dir or os.path.dirname(os.path.abspath(__file__))

def _output_path(base_dir, year, month):
    folder = os.path.join(_base(base_dir), "data", str(year), f"{month:02d}")
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, f"{month:02d}.{year}_calendar.xlsx")

def _status_path(base_dir, year, month):
    return os.path.join(_base(base_dir), "data", str(year), f"{month:02d}",
                        f"{month:02d}.{year}_status.csv")

def _filtered_path(base_dir, year, month):
    return os.path.join(_base(base_dir), "data", str(year), f"{month:02d}",
                        f"{month:02d}.{year}_filtered.xlsx")

def _tournaments_path(base_dir):
    return os.path.join(_base(base_dir), "tournaments.csv")

# ---------------------------------------------------------------------------
# Статусы дней
# ---------------------------------------------------------------------------

def _load_status_map(base_dir, year, month) -> Dict[date, str]:
    path   = _status_path(base_dir, year, month)
    result = {}
    if not os.path.exists(path):
        logger.warning(f"Status-файл не найден: {path}")
        return result
    try:
        df = pd.read_csv(path, sep=";", dtype=str)
        df.columns = [c.strip().lower() for c in df.columns]
        for _, row in df.iterrows():
            try:
                d = datetime.strptime(row["date"].strip(), "%d.%m.%Y").date()
                result[d] = STATUS_COLORS.get(row["color"].strip().lower(), DEFAULT_COLOR)
            except Exception:
                continue
        logger.info(f"Загружено {len(result)} статусов")
    except Exception as e:
        logger.exception(f"Ошибка чтения status-файла: {e}")
    return result

# ---------------------------------------------------------------------------
# Короткие названия турниров
# ---------------------------------------------------------------------------

def _load_short_names(base_dir) -> Dict[str, str]:
    path = _tournaments_path(base_dir)
    if not os.path.exists(path):
        return {}
    df = pd.read_csv(path, sep=";", dtype=str)
    df.columns = [c.strip().lower() for c in df.columns]
    if "tournament" in df.columns:
        df = df.rename(columns={"tournament": "full"})
    df["full"]  = df["full"].astype(str).str.strip()
    df["short"] = df["short"].astype(str).str.strip()
    return dict(zip(df["full"], df["short"]))


def _normalize(text: str) -> str:
    s = text.lower().replace("\xa0", " ")
    s = re.sub(r"[^a-zа-я0-9\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _get_short_name(full_name: str, short_names: Dict[str, str]) -> str:
    norm_full = _normalize(full_name)
    for key, short in short_names.items():
        if key and _normalize(key) in norm_full:
            return short
    parts = re.sub(r"[^A-Za-zА-Яа-я0-9\s]", " ", full_name).split()
    return " ".join(parts[:2]) if parts else full_name.strip()


def _build_day_tournament_line(day_df: pd.DataFrame, short_names: Dict[str, str]) -> str:
    abbrevs         = []
    champ_countries = []
    seen            = set()
    for full_name in day_df["tournament"].dropna().unique():
        short = _get_short_name(str(full_name).strip(), short_names)
        if short in seen:
            continue
        seen.add(short)
        if short.startswith("Чемпионат "):
            country = short[len("Чемпионат "):]
            if country not in champ_countries:
                champ_countries.append(country)
        else:
            if short not in abbrevs:
                abbrevs.append(short)
    parts = abbrevs[:]
    if champ_countries:
        parts.append("Чемпионаты " + ", ".join(champ_countries))
    return ", ".join(parts)

# ---------------------------------------------------------------------------
# Публичный API: карта ячеек
# ---------------------------------------------------------------------------

def get_day_cell_map(ws, year: int, month: int) -> Dict[date, Dict[str, str]]:
    """Возвращает {date: {"date_cell": "G4", "content_cell": "G5"}}."""
    result  = {}
    current = date(year, month, 1) - timedelta(days=date(year, month, 1).weekday())
    for date_row, content_row in WEEK_ROW_PAIRS:
        for dow in range(7):
            col = COL_LETTERS[dow]
            if current.month == month:
                result[current] = {
                    "date_cell":    f"{col}{date_row}",
                    "content_cell": f"{col}{content_row}",
                }
            current += timedelta(days=1)
    return result


def write_day_content(ws, cell_map, day: date, text: str):
    if day not in cell_map:
        logger.warning(f"День {day} не в cell_map")
        return
    ws[cell_map[day]["content_cell"]] = text

# ---------------------------------------------------------------------------
# Лист 2: Детально
# ---------------------------------------------------------------------------

def _fill_detail_sheet(ws2, filtered_path: str, short_names: Dict[str, str]):
    if not os.path.exists(filtered_path):
        logger.warning(f"Filtered не найден: {filtered_path}")
        return

    df = pd.read_excel(filtered_path, dtype=str, engine="openpyxl")
    df.columns = [c.strip().lower() for c in df.columns]
    df = df.fillna("")

    for col in ("tournament", "date", "time", "participants"):
        if col not in df.columns:
            df[col] = ""

    df["_st"] = df["time"].str.strip().str.replace("\xa0", "", regex=False)
    df = df.sort_values(["date", "_st"]).reset_index(drop=True)

    # Очищаем старые данные не трогая заголовок
    for row_cells in ws2.iter_rows(min_row=2):
        for cell in row_cells:
            cell.value = None

    for i, row in df.iterrows():
        r        = i + 2
        date_val = str(row["date"]).strip()
        try:
            if len(date_val) >= 10 and date_val[4] == "-":
                date_str = datetime.strptime(date_val[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
            else:
                date_str = datetime.strptime(date_val[:10], "%d.%m.%Y").strftime("%d.%m.%Y")
        except Exception:
            date_str = date_val

        ws2.cell(row=r, column=1).value = str(row["tournament"]).strip()
        ws2.cell(row=r, column=2).value = date_str
        ws2.cell(row=r, column=3).value = str(row["time"]).strip().replace("\xa0", "")
        ws2.cell(row=r, column=4).value = str(row["participants"]).strip()

    logger.info(f"Лист Детально: {len(df)} строк")

# ---------------------------------------------------------------------------
# Основная функция
# ---------------------------------------------------------------------------

def create_month_calendar(
    month_str: str,
    base_dir: Optional[str] = None,
    template_path: Optional[str] = None,
) -> str:
    """
    Создаёт Excel-календарь для месяца.

    Аргументы:
        month_str     — "03.2026"
        base_dir      — корень проекта
        template_path — путь к шаблону (по умолчанию Календарь_Событий.xlsx)

    Возвращает путь к созданному файлу.
    """
    try:
        dt = datetime.strptime(month_str, "%m.%Y")
        year, month = dt.year, dt.month
    except ValueError:
        raise ValueError(f"Неверный формат: '{month_str}'. Ожидается MM.YYYY")

    logger.info(f"Создание календаря: {month:02d}.{year}")

    if base_dir is None:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    if template_path is None:
        template_path = TEMPLATE_PATH

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")

    output_path = _output_path(base_dir, year, month)
    shutil.copy2(template_path, output_path)
    logger.info(f"Шаблон → {output_path}")

    wb = load_workbook(output_path)
    SHEET1 = "Календарь спорт. активности"
    SHEET2 = "Детально"
    ws  = wb[SHEET1] if SHEET1 in wb.sheetnames else wb.worksheets[0]
    ws2 = wb[SHEET2] if SHEET2 in wb.sheetnames else wb.create_sheet(SHEET2)

    status_map    = _load_status_map(base_dir, year, month)
    short_names   = _load_short_names(base_dir)
    filtered_path = _filtered_path(base_dir, year, month)

    # Загружаем filtered для контента ячеек
    filtered_df = None
    if os.path.exists(filtered_path):
        try:
            filtered_df = pd.read_excel(filtered_path, dtype=str, engine="openpyxl")
            filtered_df.columns = [c.strip().lower() for c in filtered_df.columns]
            filtered_df = filtered_df.fillna("")

            def _norm_date(v):
                v = str(v).strip()
                if len(v) >= 10 and v[4] == "-" and v[7] == "-":
                    return v[:10]
                for fmt in ("%d.%m.%Y", "%d/%m/%Y"):
                    try:
                        return datetime.strptime(v[:10], fmt).strftime("%Y-%m-%d")
                    except Exception:
                        pass
                return v

            filtered_df["date"] = filtered_df["date"].apply(_norm_date)
            logger.info(f"Filtered загружен: {len(filtered_df)} строк")
        except Exception as e:
            logger.exception(f"Ошибка загрузки filtered: {e}")

    # Лист 1: очистка ячеек шаблона
    for date_row, content_row in WEEK_ROW_PAIRS:
        for col in COL_LETTERS.values():
            ws[f"{col}{date_row}"].value = None
            ws[f"{col}{content_row}"].value = None

    # Лист 1: заголовок
    ws["A2"] = year
    ws["B2"] = MONTH_NAMES_RU[month]

    # Лист 1: сетка дней
    first_day = date(year, month, 1)
    current   = first_day - timedelta(days=first_day.weekday())

    for date_row, content_row in WEEK_ROW_PAIRS:
        for dow in range(7):
            col          = COL_LETTERS[dow]
            date_cell    = ws[f"{col}{date_row}"]
            content_cell = ws[f"{col}{content_row}"]

            if current.month == month:
                date_cell.value = datetime(current.year, current.month, current.day)
                fill_color      = status_map.get(current, DEFAULT_COLOR)
                fill            = PatternFill(fill_type="solid", fgColor=fill_color)
                date_cell.fill    = fill
                content_cell.fill = fill
                content_cell.value = None
                if filtered_df is not None and "tournament" in filtered_df.columns:
                    day_df = filtered_df[
                        filtered_df["date"] == current.strftime("%Y-%m-%d")
                    ]
                    if not day_df.empty:
                        content_cell.value = _build_day_tournament_line(day_df, short_names)
            else:
                date_cell.value    = None
                content_cell.value = None
                date_cell.fill     = GREY_FILL
                content_cell.fill  = GREY_FILL

            current += timedelta(days=1)

    # Лист 2: Детально
    _fill_detail_sheet(ws2, filtered_path, short_names)

    wb.save(output_path)
    logger.info(f"Сохранён → {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# Точка входа
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    month_arg = sys.argv[1] if len(sys.argv) > 1 else "06.2026"
    logger.info(f"Запуск: month={month_arg}")
    try:
        path = create_month_calendar(month_arg)
        print(f"Файл создан: {path}")
    except Exception as e:
        logger.exception(f"Ошибка: {e}")
        sys.exit(1)